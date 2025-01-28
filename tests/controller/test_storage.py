from src.anonymizer.utils.storage import (
    count_studies_series_images,
    count_series,
    get_dcm_files,
    count_study_images,
    read_java_anonymizer_index_xlsx,
    JavaAnonymizerExportedStudy,
)
import tempfile
from typing import Any, Generator, Tuple    
from openpyxl import Workbook
import pytest
from unittest.mock import patch, MagicMock
import shutil
from pathlib import Path
import os
import random


#Basic tests
def test_count_studies_series_images(temp_dir: str):
    """
    Tests the count_studies_series_images function with various scenarios.

    Assumes the path is to a patients folder?

    Args:
        temp_dir (str): The path to the temporary directory.
    """

    # Scenario 1: Empty directory
    assert count_studies_series_images(temp_dir) == (0, 0, 0)

    # Scenario 2: Single study, single series, single image
    os.makedirs(os.path.join(temp_dir, "study1", "series1"))
    with open(os.path.join(temp_dir, "study1", "series1", "image1.dcm"), "w") as _:
        pass  # Create an empty DICOM file 
    assert count_studies_series_images(temp_dir) == (1, 1, 1) 

    # Scenario 3: Multiple studies, series, and images
    os.makedirs(os.path.join(temp_dir, "study2", "series1"))
    os.makedirs(os.path.join(temp_dir, "study1", "series2"))
    with open(os.path.join(temp_dir, "study1", "series1", "image2.dcm"), "w") as _:
        pass
    with open(os.path.join(temp_dir, "study1", "series1", "image3.dcm"), "w") as _:
        pass
    with open(os.path.join(temp_dir, "study2", "series1", "image4.dcm"), "w") as _:
        pass
    assert count_studies_series_images(temp_dir) == (2, 3, 4)

    # Scenario 4: Empty series
    os.makedirs(os.path.join(temp_dir, "study3", "empty_series"))
    assert count_studies_series_images(temp_dir) == (3, 4, 4) 

    # Scenario 5: Non-DICOM files
    with open(os.path.join(temp_dir, "study1", "series1", "not_dicom.txt"), "w") as _:
        pass 
    assert count_studies_series_images(temp_dir) == (3, 4, 4)  

def test_count_series(temp_dir: str):
    """
    Tests the count_series function with various scenarios.

    Assumes the path is to the root folder for multiple patients?

    Args:
        temp_dir (str): The path to the temporary directory.
    """

    # Scenario 1: Empty directory
    assert count_series(temp_dir) == 0

    # Scenario 2: Single patient, single study, single series
    os.makedirs(os.path.join(temp_dir, "patient1", "study1", "series1"))
    assert count_series(temp_dir) == 1
    assert count_series(temp_dir, patient_ids=["patient1"]) == 1

    # Scenario 3: Multiple patients, one study each, multiple series
    os.makedirs(os.path.join(temp_dir, "patient1", "study1", "series2"))
    os.makedirs(os.path.join(temp_dir, "patient2", "study1", "series1"))
    assert count_series(temp_dir) == 3 
    assert count_series(temp_dir, patient_ids=["patient1"]) == 2

    # Scenario 4: Patient not found
    assert count_series(temp_dir, patient_ids=["non_existent_patient"]) == 0

    # Scenario 5: Empty patient directory
    os.makedirs(os.path.join(temp_dir, "empty_patient"))
    assert count_series(temp_dir) == 3
    assert count_series(temp_dir, patient_ids=["empty_patient"]) == 0 

    # Scenario 6: Invalid base directory (not a directory)
    with pytest.raises(ValueError, match="is not a valid directory"):
        count_series("/path/that/does/not/exist") 

def test_get_dcm_files(temp_dir: str):
    """
    Tests the get_dcm_files function with various scenarios.

    Args:
        temp_dir (str): The path to the temporary directory.
    """

    # Scenario 1: Empty directory
    assert get_dcm_files(temp_dir) == []

    # Scenario 2: Single DICOM file
    os.makedirs(os.path.join(temp_dir, "study1", "series1"))
    with open(os.path.join(temp_dir, "study1", "series1", "image1.dcm"), "w") as _:
        pass
    assert get_dcm_files(temp_dir) == [Path(temp_dir) / "study1" / "series1" / "image1.dcm"]

    # Scenario 3: Multiple DICOM files in different locations
    os.makedirs(os.path.join(temp_dir, "study2", "series2"))
    with open(os.path.join(temp_dir, "study1", "series1", "image2.dcm"), "w") as _:
        pass
    with open(os.path.join(temp_dir, "study2", "series2", "image3.dcm"), "w") as _:
        pass
    with open(os.path.join(temp_dir, "study2", "series2", "image4.dcm"), "w") as _:
        pass
    expected_files = [
        Path(temp_dir) / "study1" / "series1" / "image1.dcm",
        Path(temp_dir) / "study1" / "series1" / "image2.dcm",
        Path(temp_dir) / "study2" / "series2" / "image3.dcm",
        Path(temp_dir) / "study2" / "series2" / "image4.dcm",
    ]
    assert set(get_dcm_files(temp_dir)) == set(expected_files)

    # Scenario 4: Non-DICOM files
    with open(os.path.join(temp_dir, "study1", "series1", "non_dicom.txt"), "w") as _:
        pass
    assert set(get_dcm_files(temp_dir)) == set(expected_files)

    # Scenario 5: Start search at study level
    assert set(get_dcm_files(os.path.join(temp_dir, "study1"))) == set(
        [
            Path(temp_dir) / "study1" / "series1" / "image1.dcm",
            Path(temp_dir) / "study1" / "series1" / "image2.dcm",
        ]
    )

def test_count_study_images(temp_dir: str):
    """
    Tests the count_study_images function with various scenarios.

    Args:
        temp_dir (str): The path to the temporary directory.
    """
    base_dir = Path(temp_dir)

    # Scenario 1: Empty study directory
    os.makedirs(os.path.join(temp_dir, "patient1", "study1"))
    assert count_study_images(base_dir, "patient1", "study1") == 0

    # Scenario 2: Single image in study
    os.makedirs(os.path.join(temp_dir, "patient1", "study2"))
    with open(os.path.join(temp_dir, "patient1", "study2", "image1.dcm"), "w") as _:
        pass
    assert count_study_images(base_dir, "patient1", "study2") == 1

    # Scenario 3: Multiple images in study
    os.makedirs(os.path.join(temp_dir, "patient2", "study3"))
    with open(os.path.join(temp_dir, "patient2", "study3", "image1.dcm"), "w") as _:
        pass
    with open(os.path.join(temp_dir, "patient2", "study3", "image2.dcm"), "w") as _:
        pass
    assert count_study_images(base_dir, "patient2", "study3") == 2

    # Scenario 4: Non-DICOM files
    with open(os.path.join(temp_dir, "patient1", "study2", "non_dicom.txt"), "w") as _:
        pass
    assert count_study_images(base_dir, "patient1", "study2") == 1

    # Scenario 5: Study directory not found
    assert count_study_images(base_dir, "non_existent_patient", "non_existent_study") == 0


#Random number of subsiquent folders
def generate_random_dicom_structure(temp_dir: str, num_patients = 1) -> Tuple[int, int, int, int]:
    """
    Generates a random DICOM structure within the given temporary directory.

    Args:
        temp_dir: The path to the temporary directory.

    Returns:
        A tuple containing:
            - Number of patients
            - Number of studies
            - Number of series
            - Number of images
    """
    num_studies = 0
    num_series = 0
    num_images = 0

    for i in range(num_patients):
        patient_dir = os.path.join(temp_dir, f"patient_{i}")
        os.makedirs(patient_dir)

        num_patient_studies = random.randint(1, 5)
        for j in range(num_patient_studies):
            study_dir = os.path.join(patient_dir, f"study_{j}")
            os.makedirs(study_dir)

            num_patient_study_series = random.randint(1, 10)
            for k in range(num_patient_study_series):
                series_dir = os.path.join(study_dir, f"series_{k}")
                os.makedirs(series_dir)

                num_patient_study_series_images = random.randint(1, 3)
                for l in range(num_patient_study_series_images):
                    image_path = os.path.join(series_dir, f"image_{l}.dcm")
                    with open(image_path, "w") as _:
                        pass  # Create an empty DICOM file 

                    num_images += 1
                num_series += 1
            num_studies += 1

    return num_patients, num_studies, num_series, num_images

@pytest.fixture
def temp_dicom_dir_one_patient() -> Generator[Tuple[str, int, int, int, int], None, None]:
    """
    Creates a temporary directory with a random DICOM structure and yields its path 
    along with the counts of patients, studies, series, and images.

    Yields:
        A tuple containing:
            - The path to the temporary directory.
            - The number of patients.
            - The number of studies.
            - The number of series.
            - The number of images.
    """
    temp_path = tempfile.mkdtemp()
    num_patients, num_studies, num_series, num_images = generate_random_dicom_structure(temp_path, 1)
    yield temp_path, num_patients, num_studies, num_series, num_images
    shutil.rmtree(temp_path)

@pytest.fixture
def temp_dicom_dir_50_patient() -> Generator[Tuple[str, int, int, int, int], None, None]:
    """
    Creates a temporary directory with a random DICOM structure and yields its path 
    along with the counts of patients, studies, series, and images.

    Yields:
        A tuple containing:
            - The path to the temporary directory.
            - The number of patients.
            - The number of studies.
            - The number of series.
            - The number of images.
    """
    temp_path = tempfile.mkdtemp()
    num_patients, num_studies, num_series, num_images = generate_random_dicom_structure(temp_path, 50)
    yield temp_path, num_patients, num_studies, num_series, num_images
    shutil.rmtree(temp_path)

def test_count_studies_random_series_images(temp_dicom_dir_one_patient):
    """
    Tests the count_studies_series_images function with a single patient with a random number of subsiquent folders.

    Args:
        temp_dicom_dir_one_patient: A tuple containing the temporary directory path and counts.
    """
    temp_dir, expected_patients, expected_studies, expected_series, expected_images = temp_dicom_dir_one_patient
    actual_studies, actual_series, actual_images = count_studies_series_images(f"{temp_dir}/patient_0/") 
    assert actual_studies == expected_studies
    assert actual_series == expected_series
    assert actual_images == expected_images

def test_count_studies_random_count_series(temp_dicom_dir_50_patient):
    """
    Tests the count_series function with a random number of patients and subsiquent folders.

    Args:
        temp_dicom_dir_50_patient: A tuple containing the temporary directory path and counts.
    """
    temp_dir, expected_patients, expected_studies, expected_series, expected_images = temp_dicom_dir_50_patient
    actual_series = count_series(f"{temp_dir}") 
    assert actual_series == expected_series


#read_java_anonymizer_index_xlsx
def test_load_data_from_existing_excel():
    """
    Tests if the function can load data from an existing Excel file and verifies 
    the number of entries and the IDs of the first few entries.
    """
    filename = 'tests/controller/assets/JavaGeneratedIndex.xlsx' 
    data = read_java_anonymizer_index_xlsx(filename)

    # Expected number of entries (replace with actual value)
    expected_num_entries = 112  # Example

    # Assert the number of entries
    assert len(data) == expected_num_entries

    # Expected IDs for the first few entries (replace with actual values)
    expected_first_ids = ["527408-000001", "527408-000002", "527408-000003"] 

    # Assert IDs of the first few entries
    for i, expected_id in enumerate(expected_first_ids):
        assert data[i].ANON_PatientID == expected_id


def test_no_active_sheet():
    """Test that a ValueError is raised if no active sheet is found in the workbook."""
    # Mock the load_workbook function
    with patch("src.anonymizer.utils.storage.load_workbook") as mock_load_workbook:
        # Create a mock workbook with no active sheet
        mock_workbook = MagicMock()
        mock_workbook.active = None
        mock_load_workbook.return_value = mock_workbook

        # Verify the function raises ValueError
        with pytest.raises(ValueError, match="No active sheet found in the workbook"):
            read_java_anonymizer_index_xlsx("fake_filename.xlsx")
