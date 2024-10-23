"""
This module provides utility functions for working with storage in the anonymizer application.

Functions:
- count_studies_series_images(patient_path: str) -> Tuple[int, int, int]: Counts the number of studies, series, and images in a given patient directory.
- count_study_images(base_dir: Path, anon_pt_id: str, study_uid: str) -> int: Counts the number of images stored in a given study directory.
- read_java_anonymizer_index_xlsx(filename: str) -> List[JavaAnonymizerExportedStudy]: Read data from the Java Anonymizer exported patient index file.

Classes:
- JavaAnonymizerExportedStudy: Represents the data structure for a single exported study from the Java Anonymizer.

"""

import os
from typing import Union
from dataclasses import dataclass
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.child import _WorkbookChild


DICOM_FILE_SUFFIX = ".dcm"


def count_studies_series_images(patient_path: str) -> tuple[int, int, int]:
    """
    Counts the number of studies, series, and images in a given patient directory in the anonymizer store

    Args:
        patient_path (str): The path to the patient directory.

    Returns:
        A tuple containing the number of studies, series, and images in the patient directory.
    """
    study_count = 0
    series_count = 0
    image_count = 0

    for root, dirs, files in os.walk(patient_path):
        if root == patient_path:
            study_count += len(dirs)
        else:
            series_count += len(dirs)
        for file in files:
            if file.endswith(DICOM_FILE_SUFFIX):
                image_count += 1

    return study_count, series_count, image_count


def patient_dcm_files(patient_path: str) -> list[Path]:
    """
    Retrieves paths for each dicom file for a patient

    Args:
        patient_path (str): The path to the patient directory.

    Returns:
        List of Path objects
    """

    return [
        Path(root) / file
        for root, _, files in os.walk(patient_path)
        for file in files
        if file.endswith(DICOM_FILE_SUFFIX)
    ]


def count_study_images(base_dir: Path, anon_pt_id: str, study_uid: str) -> int:
    """
    Counts the number of images stored in a given study directory.

    Args:
        anon_uid (str): The anonymous patient ID.
        study_uid (str): The study UID.

    Returns:
        The number of images stored in the study directory.
    """
    study_path = Path(base_dir, anon_pt_id, study_uid)
    image_count = 0

    for _, _, files in os.walk(study_path):
        for file in files:
            if file.endswith(DICOM_FILE_SUFFIX):
                image_count += 1

    return image_count


@dataclass
class JavaAnonymizerExportedStudy:
    ANON_PatientName: str
    ANON_PatientID: str
    PHI_PatientName: str
    PHI_PatientID: str
    DateOffset: str
    ANON_StudyDate: str
    PHI_StudyDate: str
    ANON_Accession: str
    PHI_Accession: str
    ANON_StudyInstanceUID: str
    PHI_StudyInstanceUID: str


def read_java_anonymizer_index_xlsx(filename: str) -> list[JavaAnonymizerExportedStudy]:
    """
    Read data from the Java Anonymizer exported patient index file
    containing a single workbook & sheet with fields as per the JavaAnonymizerExportedStudy dataclass.

    Args:
        filename (str): The path to the Excel file.

    Returns:
        List of JavaAnonymizerExportedStudy dataclass objects.

    Raises:
        ValueError: If no active sheet is found in the workbook.
        FileNotFoundError: If the file is not found.

    If the sheet is empty, an empty list is returned.
    """

    workbook: Workbook = load_workbook(filename)
    sheet: Union[_WorkbookChild, None] = workbook.active
    data: list[JavaAnonymizerExportedStudy] = []

    if sheet is None or not isinstance(sheet, Worksheet):
        raise ValueError("No active sheet found in the workbook")

    for row in sheet.iter_rows(values_only=True, min_row=2):
        str_row = [str(item) if item is not None else "" for item in row]
        data.append(JavaAnonymizerExportedStudy(*str_row))

    return data
